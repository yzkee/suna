"""
Create and delete Excel pivot tables via LibreOffice's DataPilot engine.

Usage:
    python pivot_table.py create <excel_file> <config_json> [timeout]
    python pivot_table.py delete <excel_file> <source_sheet> <pivot_name> [timeout]

Config JSON fields:
    source_sheet   — sheet with source data (headers in row 1)
    target_sheet   — sheet for pivot output (created if missing)
    pivot_name     — unique name for the pivot table
    source_range   — optional, e.g. "A1:E100" (defaults to full used area)
    row_fields     — list of header names for row labels
    column_fields  — list of header names for column labels
    data_fields    — list of {name, function} dicts (SUM, COUNT, AVERAGE, MAX, MIN, etc.)
    page_fields    — optional list of header names for page/filter fields
"""

import json
import os
import re
import shutil
import subprocess
import sys
import time
import zipfile
from pathlib import Path

from _soffice import macro_dir, run_soffice, soffice_env

DATAPILOT_FUNCTIONS = {
    "SUM": "com.sun.star.sheet.GeneralFunction.SUM",
    "COUNT": "com.sun.star.sheet.GeneralFunction.COUNT",
    "AVERAGE": "com.sun.star.sheet.GeneralFunction.AVERAGE",
    "MAX": "com.sun.star.sheet.GeneralFunction.MAX",
    "MIN": "com.sun.star.sheet.GeneralFunction.MIN",
    "PRODUCT": "com.sun.star.sheet.GeneralFunction.PRODUCT",
    "STDEV": "com.sun.star.sheet.GeneralFunction.STDEV",
    "STDEVP": "com.sun.star.sheet.GeneralFunction.STDEVP",
    "VAR": "com.sun.star.sheet.GeneralFunction.VAR",
    "VARP": "com.sun.star.sheet.GeneralFunction.VARP",
}

SCRIPT_XLB_CONTENT = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<!DOCTYPE library:library PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "library.dtd">\n'
    '<library:library xmlns:library="http://openoffice.org/2000/library" '
    'library:name="Standard" library:readonly="false" library:passwordprotected="false">\n'
    ' <library:element library:name="Module1"/>\n'
    "</library:library>"
)

DEFAULT_TIMEOUT = 20


def _esc(s: str) -> str:
    return s.replace('"', '""').replace("\n", " ").replace("\r", " ")


def _parse_cell_ref(ref: str) -> tuple[int, int]:
    col = 0
    row_str = ""
    for ch in ref:
        if ch.isalpha():
            col = col * 26 + (ord(ch.upper()) - ord("A") + 1)
        else:
            row_str += ch
    return int(row_str) - 1, col - 1


def _install_macro(content: str) -> bool:
    mdir = macro_dir()
    macro_file = mdir / "Module1.xba"

    if not mdir.exists():
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=soffice_env(),
        )
        time.sleep(1)
        mdir.mkdir(parents=True, exist_ok=True)

    script_xlb = mdir / "script.xlb"
    if not script_xlb.exists():
        script_xlb.write_text(SCRIPT_XLB_CONTENT)

    try:
        macro_file.write_text(content)
        return True
    except Exception:
        return False


def _run_macro(path: str, macro_name: str, timeout: int = DEFAULT_TIMEOUT) -> subprocess.CompletedProcess[str]:
    uri = f"vnd.sun.star.script:Standard.Module1.{macro_name}?language=Basic&location=application"
    return run_soffice(["--headless", "--norestore", uri, path], timeout=timeout)


def _count_pivots(path: str) -> int:
    try:
        with zipfile.ZipFile(path) as zf:
            return sum(1 for n in zf.namelist() if "pivotTable" in n and n.endswith(".xml"))
    except Exception:
        return 0


def _fix_multi_data_pivots(path: str) -> None:
    tmp = path + ".tmp"
    modified = False
    with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(tmp, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/pivotTables/pivotTable") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                df_match = re.search(r'<dataFields count="(\d+)"', text)
                if df_match and int(df_match.group(1)) > 1 and "<colFields" not in text:
                    text = text.replace(
                        "</dataFields>", '</dataFields><colFields count="1"><field x="-2"/></colFields>'
                    )
                    data = text.encode("utf-8")
                    modified = True
            zout.writestr(item, data)

    if modified:
        shutil.move(tmp, path)
    else:
        os.remove(tmp)


def _build_field_block(config: dict) -> str:
    row_fields = config.get("row_fields", [])
    column_fields = config.get("column_fields", [])
    data_fields = config.get("data_fields", [])
    page_fields = config.get("page_fields", [])

    seen_data: set[str] = set()
    cases = []

    for f in row_fields:
        cases.append(
            f'        If oField.Name = "{_esc(f)}" Then\n'
            f"            oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.ROW\n"
            f"        End If"
        )
    for f in column_fields:
        cases.append(
            f'        If oField.Name = "{_esc(f)}" Then\n'
            f"            oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.COLUMN\n"
            f"        End If"
        )
    for df in data_fields:
        name = df["name"]
        if name in seen_data:
            continue
        seen_data.add(name)
        func = DATAPILOT_FUNCTIONS.get(df["function"].upper(), DATAPILOT_FUNCTIONS["SUM"])
        cases.append(
            f'        If oField.Name = "{_esc(name)}" Then\n'
            f"            oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.DATA\n"
            f"            oField.Function = {func}\n"
            f"        End If"
        )
    for f in page_fields:
        cases.append(
            f'        If oField.Name = "{_esc(f)}" Then\n'
            f"            oField.Orientation = com.sun.star.sheet.DataPilotFieldOrientation.PAGE\n"
            f"        End If"
        )

    return "\n".join(cases)


def _build_source_range_code(source_range: str) -> str:
    if source_range:
        parts = source_range.replace("$", "").split(":")
        start_row, start_col = _parse_cell_ref(parts[0])
        end_row, end_col = _parse_cell_ref(parts[1])
        return (
            f"    Dim oSourceRange As New com.sun.star.table.CellRangeAddress\n"
            f"    oSourceRange.Sheet = oDataSheet.getRangeAddress().Sheet\n"
            f"    oSourceRange.StartColumn = {start_col}\n"
            f"    oSourceRange.StartRow = {start_row}\n"
            f"    oSourceRange.EndColumn = {end_col}\n"
            f"    oSourceRange.EndRow = {end_row}\n"
            f"    oDPDesc.setSourceRange(oSourceRange)"
        )
    return (
        "    Dim oCursor As Object\n"
        "    oCursor = oDataSheet.createCursor()\n"
        "    oCursor.gotoStartOfUsedArea(False)\n"
        "    oCursor.gotoEndOfUsedArea(True)\n"
        "    oDPDesc.setSourceRange(oCursor.getRangeAddress())"
    )


def _generate_macro(config: dict) -> str:
    source_sheet = config["source_sheet"]
    target_sheet = config["target_sheet"]
    pivot_name = config.get("pivot_name", "Pivot1")

    field_block = _build_field_block(config)
    source_range_code = _build_source_range_code(config.get("source_range", ""))

    return f'''\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" \
script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave()
    ThisComponent.calculateAll()
    ThisComponent.store()
    ThisComponent.close(True)
End Sub

Sub CreatePivotTable()

    Dim oDoc As Object
    oDoc = ThisComponent
    Dim oSheets As Object
    oSheets = oDoc.Sheets
    Dim oDataSheet As Object
    oDataSheet = oSheets.getByName("{_esc(source_sheet)}")

    If Not oSheets.hasByName("{_esc(target_sheet)}") Then
        oSheets.insertNewByName("{_esc(target_sheet)}", oSheets.Count)
    End If

    Dim oPivotSheet As Object
    oPivotSheet = oSheets.getByName("{_esc(target_sheet)}")
    Dim oTarget As New com.sun.star.table.CellAddress
    oTarget.Sheet = oPivotSheet.getRangeAddress().Sheet
    oTarget.Column = 0
    oTarget.Row = 0

    Dim oDPTables As Object
    oDPTables = oDataSheet.getDataPilotTables()
    Dim oDPDesc As Object
    oDPDesc = oDPTables.createDataPilotDescriptor()

{source_range_code}

    Dim oFields As Object
    oFields = oDPDesc.getDataPilotFields()
    Dim i As Long
    For i = 0 To oFields.Count - 1
        Dim oField As Object
        oField = oFields.getByIndex(i)
{field_block}
    Next i

    oDPTables.insertNewByName("{_esc(pivot_name)}", oTarget, oDPDesc)

    oDoc.store()
    oDoc.close(True)
End Sub
</script:module>'''


def _generate_delete_macro(source_sheet: str, pivot_name: str) -> str:
    return f'''\
<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" \
script:name="Module1" script:language="StarBasic">
Sub RecalculateAndSave()
    ThisComponent.calculateAll()
    ThisComponent.store()
    ThisComponent.close(True)
End Sub

Sub DeletePivotTable()
    Dim oDoc As Object
    oDoc = ThisComponent
    Dim oSheet As Object
    oSheet = oDoc.Sheets.getByName("{_esc(source_sheet)}")
    Dim oDPTables As Object
    oDPTables = oSheet.getDataPilotTables()

    Dim i As Long
    For i = 0 To oDPTables.Count - 1
        If oDPTables.getByIndex(i).Name = "{_esc(pivot_name)}" Then
            oDPTables.removeByName("{_esc(pivot_name)}")
            Exit For
        End If
    Next i

    oDoc.store()
    oDoc.close(True)
End Sub
</script:module>'''


def _validate_config(filename: str, config: dict) -> str | None:
    from openpyxl import load_workbook

    source_sheet = config.get("source_sheet", "")
    if not source_sheet:
        return "source_sheet is required"

    if not config.get("target_sheet"):
        return "target_sheet is required"

    if not config.get("data_fields"):
        return "data_fields is required (at least one)"

    data_names = [df["name"] for df in config["data_fields"]]
    if len(data_names) != len(set(data_names)):
        return "Multiple aggregations on the same field are not supported. Use separate pivot tables instead."

    try:
        wb = load_workbook(filename, read_only=True, data_only=True)
    except Exception as e:
        return f"Cannot open file: {e}"

    try:
        if source_sheet not in wb.sheetnames:
            return f"Source sheet '{source_sheet}' not found. Available: {wb.sheetnames}"

        ws = wb[source_sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        all_fields = (
            config.get("row_fields", [])
            + config.get("column_fields", [])
            + [df["name"] for df in config.get("data_fields", [])]
            + config.get("page_fields", [])
        )
        missing = [f for f in all_fields if f not in headers]
        if missing:
            return f"Fields not found in headers: {missing}. Available: {headers}"
    finally:
        wb.close()

    return None


def create_pivot(filename: str, config: dict, timeout: int = DEFAULT_TIMEOUT) -> dict:
    path = Path(filename)
    if not path.exists():
        return {"status": "error", "error": f"File {filename} does not exist"}

    abs_path = str(path.absolute())
    pivot_name = config.get("pivot_name", "Pivot1")

    error = _validate_config(abs_path, config)
    if error:
        return {"status": "error", "error": error}

    pivots_before = _count_pivots(abs_path)

    if not _install_macro(_generate_macro(config)):
        return {"status": "error", "error": "Failed to install LibreOffice macro"}

    result = _run_macro(abs_path, "CreatePivotTable", timeout)
    if result.returncode not in (0, 124):
        return {"status": "error", "error": result.stderr or "Unknown error creating pivot table"}

    pivots_after = _count_pivots(abs_path)
    if pivots_after <= pivots_before:
        return {
            "status": "error",
            "error": f"Pivot table '{pivot_name}' was not created. Check that field names match column headers exactly.",
        }

    if len(config.get("data_fields", [])) > 1:
        _fix_multi_data_pivots(abs_path)

    return {
        "status": "success",
        "pivot_tables": pivots_after,
        "target_sheet": config["target_sheet"],
    }


def delete_pivot(filename: str, source_sheet: str, pivot_name: str, timeout: int = DEFAULT_TIMEOUT) -> dict:
    path = Path(filename)
    if not path.exists():
        return {"status": "error", "error": f"File {filename} does not exist"}

    abs_path = str(path.absolute())

    if not _install_macro(_generate_delete_macro(source_sheet, pivot_name)):
        return {"status": "error", "error": "Failed to install LibreOffice macro"}

    result = _run_macro(abs_path, "DeletePivotTable", timeout)
    if result.returncode not in (0, 124):
        return {"status": "error", "error": result.stderr or "Unknown error deleting pivot table"}

    return {"status": "success", "deleted": pivot_name}


def main() -> None:
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1]

    if command == "create":
        if len(sys.argv) < 4:
            print("Usage: pivot_table.py create <excel_file> <config_json> [timeout]")
            sys.exit(1)
        filename = sys.argv[2]
        config = json.loads(sys.argv[3])
        timeout = int(sys.argv[4]) if len(sys.argv) > 4 else DEFAULT_TIMEOUT
        print(json.dumps(create_pivot(filename, config, timeout=timeout), indent=2))

    elif command == "delete":
        if len(sys.argv) < 5:
            print("Usage: pivot_table.py delete <excel_file> <source_sheet> <pivot_name> [timeout]")
            sys.exit(1)
        filename = sys.argv[2]
        source_sheet = sys.argv[3]
        pivot_name = sys.argv[4]
        timeout = int(sys.argv[5]) if len(sys.argv) > 5 else DEFAULT_TIMEOUT
        print(json.dumps(delete_pivot(filename, source_sheet, pivot_name, timeout=timeout), indent=2))

    else:
        filename = command
        config = json.loads(sys.argv[2])
        timeout = int(sys.argv[3]) if len(sys.argv) > 3 else DEFAULT_TIMEOUT
        print(json.dumps(create_pivot(filename, config, timeout=timeout), indent=2))


if __name__ == "__main__":
    main()
