from core.agentpress.tool import tool_metadata
from core.sandbox.tool_base import SandboxToolsBase
from core.agentpress.thread_manager import ThreadManager

@tool_metadata(
    display_name="Spreadsheet Guide",
    description="Reference guide for creating and manipulating Excel spreadsheets via CLI Python",
    icon="Table",
    color="bg-green-100 dark:bg-green-800/50",
    weight=75,
    visible=True,
    usage_guide="""
## 🗣️ COMMUNICATION RULES - HOW TO TALK TO USERS

**THE USER IS NON-TECHNICAL. NEVER expose implementation details.**

**DO:**
- ✅ "I'll create that spreadsheet for you"
- ✅ "Here's your budget spreadsheet with the calculations"
- ✅ "I've organized the data into a spreadsheet"
- ✅ "I've added a new sheet for Q2 data"
- ✅ "The totals are automatically calculated"

**DON'T:**
- ❌ "I'll use openpyxl to create an .xlsx file"
- ❌ "I'm executing a Python script via execute_command"
- ❌ "I'll load_workbook and update cells"
- ❌ "I'm creating a temporary Python file"
- ❌ "I'll use PatternFill and Font classes"

**TONE:**
- Friendly and conversational
- Focus on WHAT you're creating, not HOW you're building it
- Describe the outcome, not the technical process
- Make it feel effortless and natural

**EXAMPLES:**

User: "Can you create a budget spreadsheet?"
❌ BAD: "I'll create a Python script using openpyxl to generate an .xlsx file with formulas."
✅ GOOD: "I'll create a budget spreadsheet for you with all the formulas set up!"

User: "Add a column for expenses"
❌ BAD: "I'll load the workbook and insert a new column using ws.insert_cols()."
✅ GOOD: "I'll add an expenses column for you!"

User: "Can you make the headers prettier?"
❌ BAD: "I'll apply PatternFill with #1F4E79 and Font with color FFFFFF."
✅ GOOD: "I'll style the headers to make them look nicer!"

---

### SPREADSHEET OPERATIONS - Using CLI Python via execute_command

**PURPOSE:**
Create interactive Excel (.xlsx) files that users can view, edit, and download. Uses native Excel format with full formula support, formatting, and multi-sheet capabilities.

**WHEN TO USE:**
- User asks to create/organize data in spreadsheet format
- Data needs calculations, formulas, or structured presentation
- User wants downloadable Excel files
- Data visualization with formatting and colors

**HOW TO USE:**
Create temporary Python files and execute them using `execute_command`. This approach is cleaner and easier to debug than inline Python code.
Save spreadsheets anywhere in `/workspace/` - use logical paths based on context (e.g., `/workspace/reports/budget.xlsx`, `/workspace/data/analysis.xlsx`, or `/workspace/spreadsheets/` if organizing multiple spreadsheets).

**PATTERN:**
1. Create a temporary Python file using `create_file` with a `.py` extension
2. Execute it using `execute_command` with `python3 /path/to/temp_file.py`
3. Optionally clean up the temp file after execution
4. **🚨 MANDATORY FINAL STEP:** Use `ask` or `complete` with the spreadsheet in `attachments`

**FEATURES:**
- Native Excel (.xlsx) format - fully compatible
- Formulas (=SUM, =AVERAGE, =IF, etc.)
- Cell formatting (colors, fonts, alignment)
- Auto-styled headers (dark blue with white text)
- Interactive viewing in chat interface
- Downloadable Excel files
- Multi-sheet workbooks

---

## OPERATIONS

### 1. CREATE NEW SPREADSHEET
Use `openpyxl.Workbook()` to create a new workbook:

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_create_spreadsheet.py",
    file_contents='''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Q1 Sales"

# Headers with styling
headers = ["Product", "Revenue", "Profit", "Margin %"]
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='left', vertical='center')

# Data rows (starting at row 2)
data = [
    ["Product A", 50000, 15000, "=IFERROR(C2/B2*100,0)"],
    ["Product B", 75000, 22500, "=IFERROR(C3/B3*100,0)"],
    ["Total", "=SUM(B2:B3)", "=SUM(C2:C3)", "=IFERROR(C4/B4*100,0)"]
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

# Auto-width columns
for col in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 15

wb.save("/workspace/sales_report.xlsx")
print("Created: /workspace/sales_report.xlsx")'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_create_spreadsheet.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_create_spreadsheet.py")
```

**Step 4: 🚨 MANDATORY - Deliver to user**
```
complete(
    text="Here's your sales report! I've added formulas to calculate totals and margins automatically.",
    attachments="/workspace/sales_report.xlsx"
)
```
OR if you need user feedback:
```
ask(
    text="I've created your sales report with automatic calculations. Would you like me to adjust anything?",
    attachments="/workspace/sales_report.xlsx"
)
```

**Remember: Talk about RESULTS, not implementation. Focus on what the user gets, not how you built it.**

### 2. ADD SHEET TO EXISTING FILE
Use `wb.create_sheet()` to add sheets - this preserves existing sheets:

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_add_sheet.py",
    file_contents='''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook("/workspace/sales_report.xlsx")
ws = wb.create_sheet(title="Q2 Sales")

headers = ["Product", "Revenue", "Profit"]
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

data = [
    ["Product A", 60000, 18000],
    ["Product B", 80000, 24000]
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

for col in range(1, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col)].width = 15

wb.save("/workspace/sales_report.xlsx")
print("Added sheet: Q2 Sales")'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_add_sheet.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_add_sheet.py")
```

### 3. UPDATE EXISTING CELLS
Use `load_workbook()` and modify cells directly:

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_update_cells.py",
    file_contents='''import openpyxl

wb = openpyxl.load_workbook("/workspace/sales_report.xlsx")
ws = wb.active  # or wb["Sheet Name"]

# Update specific cells
ws["A1"] = "Product Name"
ws["B5"] = 100000
ws["C5"] = "=B5*0.3"

wb.save("/workspace/sales_report.xlsx")
print("Updated cells")'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_update_cells.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_update_cells.py")
```

### 4. READ SPREADSHEET DATA

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_read_spreadsheet.py",
    file_contents='''import openpyxl

wb = openpyxl.load_workbook("/workspace/sales_report.xlsx")
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Sheets in workbook: {wb.sheetnames}")

for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    print(row)'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_read_spreadsheet.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_read_spreadsheet.py")
```

### 5. FORMAT CELLS

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_format_cells.py",
    file_contents='''import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook("/workspace/sales_report.xlsx")
ws = wb.active

# Apply formatting to a range
for row in ws["A1:D1"]:
    for cell in row:
        cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.alignment = Alignment(horizontal='center')

# Add borders
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws["A1:D10"]:
    for cell in row:
        cell.border = thin_border

wb.save("/workspace/sales_report.xlsx")
print("Formatted")'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_format_cells.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_format_cells.py")
```

### 6. ADD ROWS TO EXISTING SHEET

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_add_rows.py",
    file_contents='''import openpyxl

wb = openpyxl.load_workbook("/workspace/sales_report.xlsx")
ws = wb.active

# Append new rows at the end
ws.append(["Product C", 60000, 18000, 30])
ws.append(["Product D", 45000, 13500, 30])

wb.save("/workspace/sales_report.xlsx")
print("Added rows")'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_add_rows.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_add_rows.py")
```

### 7. CROSS-SHEET REFERENCES

**Step 1: Create temporary Python file**
```
create_file(
    file_path="temp_cross_sheet.py",
    file_contents='''import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.load_workbook("/workspace/sales_report.xlsx")

# Create summary sheet
ws = wb.create_sheet(title="Summary")

headers = ["Metric", "Value"]
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF')

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font

# Reference other sheets
data = [
    ["Q1 Total Revenue", "=SUM('Q1 Sales'!B2:B10)"],
    ["Q2 Total Revenue", "=SUM('Q2 Sales'!B2:B10)"],
    ["Combined Total", "=B2+B3"]
]

for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save("/workspace/sales_report.xlsx")
print("Created summary with cross-sheet references")'''
)
```

**Step 2: Execute the Python file**
```
execute_command(
    command="python3 temp_cross_sheet.py",
    blocking=true
)
```

**Step 3: Clean up (optional)**
```
delete_file(file_path="temp_cross_sheet.py")
```

---

## IMPORTANT: DO NOT CREATE CORRUPT SPREADSHEETS

### CRITICAL - AVOIDING CIRCULAR REFERENCES

**ROW NUMBERING:**
Headers are in ROW 1. Data rows start at ROW 2.
- Row index 0 in your data array = Excel Row 2
- Row index 1 in your data array = Excel Row 3
- etc.

**CORRECT Example (headers in row 1, data starts row 2):**
If you have 3 data rows (rows 2-4), a SUM formula in row 5 should be: `=SUM(B2:B4)`
NEVER reference the cell you're placing the formula in.
Total row formulas should reference ONLY the data rows ABOVE them.

```python
# CORRECT: 3 products + 1 total row
headers = ["Product", "Revenue", "Profit"]
data = [
    ["Product A", 50000, 15000],      # Row 2
    ["Product B", 75000, 22500],      # Row 3  
    ["Product C", 60000, 18000],      # Row 4
    ["Total", "=SUM(B2:B4)", "=SUM(C2:C4)"]  # Row 5 - formulas sum rows 2-4
]
```

**WRONG (causes circular reference):**
```python
# WRONG: This creates circular reference!
data = [
    ["Product A", 50000, 15000],
    ["Product B", 75000, 22500],
    ["Product C", 60000, 18000],
    ["Total", "=SUM(B2:B5)", "=SUM(C2:C5)"]  # BAD: B5/C5 is THIS row!
]
```

### CRITICAL - PREVENTING #DIV/0! ERRORS

ALWAYS wrap division formulas with IFERROR to handle divide-by-zero:
- Use: `=IFERROR(A1/B1,0)` instead of `=A1/B1`
- For percentage: `=IFERROR(C4/B4*100,0)` instead of `=C4/B4*100`

```python
# CORRECT: Safe division with IFERROR
data = [
    ["Product A", 50000, 35000, "=IFERROR(C2/B2*100,0)"],
    ["Product B", 75000, 45000, "=IFERROR(C3/B3*100,0)"],
    ["Total", "=SUM(B2:B3)", "=SUM(C2:C3)", "=IFERROR(C4/B4*100,0)"]
]
```

---

## BEST PRACTICES

1. Count your data rows carefully before writing formulas
2. Total/summary formulas should be in the LAST row
3. Formula ranges should end at the row BEFORE the formula row
4. Use explicit values instead of formulas when unsure
5. Test mentally: "Does this formula reference its own cell?" → If yes, FIX IT
6. ALWAYS use IFERROR() around any division to prevent #DIV/0! errors
7. Use `wb.create_sheet()` to add sheets - NEVER recreate the workbook
8. **🚨 CRITICAL:** ALWAYS use `ask` or `complete` with spreadsheet in `attachments` - user CANNOT see it otherwise!

---

## COMMON FORMULAS

- SUM: `=SUM(A1:A10)`
- AVERAGE: `=AVERAGE(B2:B10)`
- COUNT: `=COUNT(A1:A100)`
- IF: `=IF(A1>100,"High","Low")`
- VLOOKUP: `=VLOOKUP(A1,Sheet2!A:B,2,FALSE)`
- SUMIF: `=SUMIF(A:A,"Product A",B:B)`
- Safe division: `=IFERROR(C2/B2*100,0)`
- Cross-sheet: `=SUM('Sheet Name'!B2:B10)`

---

## FORMATTING OPTIONS

**Colors (hex without #):**
- 4CAF50 (green)
- F44336 (red)
- 2196F3 (blue)
- FFC107 (yellow)
- 1F4E79 (dark blue - default header)
- FFFFFF (white)
- 000000 (black)

**Styles:**
- PatternFill(start_color='HEX', end_color='HEX', fill_type='solid')
- Font(bold=True, italic=True, color='HEX', size=12, name='Arial')
- Alignment(horizontal='center', vertical='center', wrap_text=True)
- Border with Side(style='thin'/'medium'/'thick')

---

## FILE PATHS

- Save spreadsheets anywhere in `/workspace/` - use logical paths based on context
- Examples:
  - `/workspace/budget.xlsx` - simple file in workspace root
  - `/workspace/reports/q1_sales.xlsx` - organized in subdirectory
  - `/workspace/data/analysis.xlsx` - data directory
  - `/workspace/spreadsheets/` - optional directory for organizing multiple spreadsheets
- Use `load_workbook()` to modify existing files
- Use `Workbook()` only for NEW files (overwrites existing!)
- Create directories if needed: `import os; os.makedirs('/workspace/reports', exist_ok=True)`

---

## 🚨🚨🚨 MANDATORY FINAL STEP: DELIVER SPREADSHEET TO USER 🚨🚨🚨

**AFTER CREATING/UPDATING ANY SPREADSHEET, YOU MUST USE `ask` OR `complete` WITH ATTACHMENT:**

```
complete(
    text="Here's your [spreadsheet name]! I've set up [what it does in user-friendly terms].",
    attachments="/workspace/your_spreadsheet.xlsx"
)
```
OR if you need user feedback:
```
ask(
    text="I've created your [spreadsheet name] with [features]. Would you like me to adjust anything?",
    attachments="/workspace/your_spreadsheet.xlsx"
)
```

**COMMUNICATION REMINDERS:**
- ✅ Describe what the spreadsheet DOES, not how you built it
- ✅ Use friendly, conversational language
- ✅ Mention key features users will care about (calculations, formatting, organization)
- ❌ NEVER mention technical details (openpyxl, Python, execute_command, etc.)
- ❌ NEVER say "I created the spreadsheet at /workspace/file.xlsx" without attaching it
- ❌ NEVER end task without using `ask` or `complete` with the spreadsheet in `attachments`

**THIS IS NOT OPTIONAL - IT IS MANDATORY:**
- Users see deliverables ONLY through `ask`/`complete` attachments parameter
- Without attachment, the spreadsheet is INVISIBLE to the user
- This breaks the entire user experience - user paid for a deliverable they can't see

**GOOD EXAMPLES:**
- "Here's your budget spreadsheet! All the totals calculate automatically, and I've color-coded the sections for easy reading."
- "I've created your sales report with quarterly breakdowns. The formulas will update automatically when you add new data."
- "Your inventory spreadsheet is ready! I've organized it by category with automatic stock calculations."
"""
)
class SandboxSpreadsheetTool(SandboxToolsBase):
    """
    Guide-only tool for spreadsheet operations.
    Use create_file + execute_command with Python/openpyxl scripts - see usage_guide for examples.
    Creates temporary Python files instead of using inline Python code for better maintainability.
    """

    def __init__(self, project_id: str, thread_manager: ThreadManager):
        super().__init__(project_id, thread_manager)
